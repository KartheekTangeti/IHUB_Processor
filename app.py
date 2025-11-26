
import time
import uuid
import tempfile
import re
from pathlib import Path
from flask import (
    Flask, request, render_template, send_file, jsonify, abort,
    url_for, after_this_request
)
from werkzeug.utils import secure_filename
import openpyxl
from lxml import etree

APP_HOST = "127.0.0.1"
APP_PORT = 8000

app = Flask(__name__)
app.config["MAX_CONTENT_LENGTH"] = 64 * 1024 * 1024  # 64 MB
ALLOWED_EXTENSIONS = {".xlsx"}
PROCESS_CACHE = {}

# Robust closing tag detector (handles namespaces and whitespace, case-insensitive)
CLOSE_TAG_RE = re.compile(r'</\s*(?:\w+:)?intercompanymessage\s*>', re.IGNORECASE)
OPEN_TAG_RE  = re.compile(r'<\s*(?:\w+:)?intercompanymessage\b', re.IGNORECASE)


def _is_allowed(filename: str) -> bool:
    return Path(filename).suffix.lower() in ALLOWED_EXTENSIONS


def _extract_messages(raw: str) -> list[str]:
    # Clean BOM, trim
    raw = raw.replace("\ufeff", "").strip()
    # Extract <intercompanyMessage>...</intercompanyMessage> blocks (case-insensitive, optional namespace)
    pattern = re.compile(
        r'(<\s*(?:\w+:)?intercompanymessage[\s\S]*?</\s*(?:\w+:)?intercompanymessage\s*>)',
        re.IGNORECASE
    )
    matches = pattern.findall(raw)
    if not matches:
        raise ValueError("No <intercompanyMessage> blocks found.")
    cleaned = []
    for msg in matches:
        # Remove optional XML declarations inside the chunk
        msg = re.sub(r'\s*<\?xml[^\>]*\?>', '', msg, flags=re.IGNORECASE)
        cleaned.append(msg.strip())
    return cleaned


def process_xml_and_populate_xl_sheet(xml_content: str, xl_sheet, current_row: int) -> int:
    # Tolerant XML parsing
    parser = etree.XMLParser(recover=True)
    root = etree.fromstring(xml_content, parser=parser)

    # Helper: safe single value
    def x1(path: str, default: str = "") -> str:
        r = root.xpath(path)
        return r[0] if r else default

    # Helper: list per line item
    LINE_ITEMS = root.xpath('//purchaseOrder//lineItems//lineItem')

    def get_data(locator: str) -> list[str]:
        return [" ".join(li.xpath(locator)) for li in LINE_ITEMS]

    # Header/scalar fields (safe)
    PUSB = x1('//purchaseOrder/@PUSB')
    PO_NUMBER = x1('//purchaseOrder/@orderNumber')
    SOS = x1('//purchaseOrder//header//SoS/text()')
    CUSTPROFCODE = x1('//purchaseOrder//header//customerProfileCode/text()')
    if CUSTPROFCODE != "":
        CUSTPROFCODE = " ".join(["STC", CUSTPROFCODE])

    ITRANSPROUTECODE = x1('//purchaseOrder//header//internationalTransportationRouteCode/text()')

    # Dates (safe parsing)
    from datetime import datetime
    po_create_raw = x1('//purchaseOrder//header//purchaseOrderCreationDate/text()')
    try:
        POCREATEDATE = datetime.strptime(po_create_raw, "%Y-%m-%d").strftime("%d.%m.%Y") if po_create_raw else ""
    except Exception:
        POCREATEDATE = ""

    # Line-level fields
    seqs_raw = get_data('./@sequenceNumber')
    POLINESEQNRS = []
    for i in seqs_raw:
        i = (i or "").strip()
        if i.isdigit():
            POLINESEQNRS.append(str(int(i)))
        else:
            POLINESEQNRS.append(i)

    MMMPRODID = get_data('./productIdentifier/text()')
    ORDERQTY = get_data('./orderQuantity/text()')
    SELLINGUNIT = get_data('./sellingUnit/text()')
    SUPPLY_CHAIN_UNIT = SELLINGUNIT

    PRODUCT_DESCRIPTION = get_data('./lineItemDetails/lineItemDetail[@type="purchaseritemdescription"]/text()')
    SPECIAL_HANDLING = get_data('./lineItemDetails/lineItemDetail[@type="specialhandlingcode"]/text()')

    no_line_items = len(POLINESEQNRS)
    LINE_INSTRUCTION = [""] * no_line_items

    ADDRESS = "; ".join(
        root.xpath('//purchaseOrder//header//purchaseOrderDetails//purchaseOrderDetail[@type="shiptoaddress"]/text()')[::-1]
    )

    EXPORT_MARKS = [""] * no_line_items

    ORDER_INSTRUCTION = "".join(
        root.xpath('//purchaseOrder//header//specialInstructions//specialInstruction[@type="AH"]/text()')
    )
    if ORDER_INSTRUCTION != "":
        ORDER_INSTRUCTION = "C" + ORDER_INSTRUCTION
    else:
        ORDER_INSTRUCTION = "null"

    EXPC_SHIP_TYPE_CODE = get_data('./requestedShipmentDate/@type')

    EXPC_SHIP_DATE = []
    for date in get_data('./requestedShipmentDate/text()'):
        try:
            if str(date).strip() not in [None, "None", ""]:
                EXPC_SHIP_DATE.append(datetime.strptime(date, "%Y-%m-%d").strftime("%d.%m.%Y"))
            else:
                EXPC_SHIP_DATE.append("")
        except Exception:
            EXPC_SHIP_DATE.append("")

    SAP_PO_NUMBER = get_data('./purchasingCompanyReferenceNumber/text()')
    SAP_PO_NUMBER = ["null" if (po_number or "") == "" else po_number for po_number in SAP_PO_NUMBER]

    # Pack all columns in order
    data = [
        PUSB, PO_NUMBER, SOS, CUSTPROFCODE, ITRANSPROUTECODE, POCREATEDATE,
        POLINESEQNRS, MMMPRODID, ORDERQTY, SELLINGUNIT, SUPPLY_CHAIN_UNIT,
        PRODUCT_DESCRIPTION, SPECIAL_HANDLING, LINE_INSTRUCTION, ADDRESS,
        EXPORT_MARKS, ORDER_INSTRUCTION, EXPC_SHIP_TYPE_CODE, EXPC_SHIP_DATE,
        SAP_PO_NUMBER
    ]

    # Write one row per line item
    for cur_record_num in range(no_line_items):
        for col in range(len(data)):
            cur_val = data[col]
            if isinstance(cur_val, list):
                # Per-line-item column
                v = cur_val[cur_record_num] if cur_record_num < len(cur_val) else ""
                xl_sheet.cell(row=current_row, column=col + 1).value = str(v)
            else:
                # Scalar column (same for all line items)
                xl_sheet.cell(row=current_row, column=col + 1).value = str(cur_val)
        # Advance row ONCE per record
        current_row += 1

    return current_row


def process_excel(xl_data_file: str, xl_output_file: str) -> str:
    MAX_PARTS = 1000
    MAX_XML_BYTES = 8 * 1024 * 1024

    wb_in = openpyxl.load_workbook(xl_data_file, read_only=True, data_only=True)
    ws_in = wb_in.active
    colA = ["" if r[0] is None else str(r[0]) for r in ws_in.iter_rows(min_row=1, max_col=1, max_row=ws_in.max_row, values_only=True)]

    wb_out = openpyxl.Workbook()
    ws_out = wb_out.active

    headers = [
        "PUSB", "PO_NUMBER", "SOS", "CUSTPROFCODE", "ITRANSPROUTECODE", "POCREATEDATE",
        "POLINESEQNR", "MMMPRODID", "ORDERQTY", "SELLINGUNIT", "SUPPLY CHAIN UNIT",
        "PRODUCT DESCRIPTION", "SPECIAL HANDLING", "LINE INSTRUCTION", "ADDRESS",
        "EXPORT MARKS", "ORDER INSTRUCTION", "EXPC SHIP TYPE CODE", "EXPC SHIP DATE",
        "SAP PO NUMBER"
    ]

    current_row = 1
    for col, val in enumerate(headers, start=1):
        ws_out.cell(row=current_row, column=col).value = val
    current_row += 1

    n = len(colA)
    i = 0
    while i < n:
        cell = colA[i] or ""
        has_open = OPEN_TAG_RE.search(cell) is not None
        has_close = CLOSE_TAG_RE.search(cell) is not None

        if has_open or ("<?xml" in cell):
            # --- Fast path: the entire message is within the SAME cell ---
            if has_open and has_close:
                raw_xml = cell
                if len(raw_xml.encode("utf-8")) > MAX_XML_BYTES:
                    app.logger.warning(f"Single-cell chunk at row {i+1} exceeds size limit. Skipping...")
                    i += 1
                    continue
                try:
                    messages = _extract_messages(raw_xml)
                except Exception as e:
                    app.logger.warning(f"Message extraction failed at row {i+1}: {e}")
                    i += 1
                    continue
                for msg in messages:
                    current_row = process_xml_and_populate_xl_sheet(msg, ws_out, current_row)
                i += 1
                continue

            # --- Multi-row path: scan forward to find the closing tag on subsequent rows ---
            parts = [cell]
            j = i + 1
            while j < n and not CLOSE_TAG_RE.search(colA[j] or ""):
                parts.append(colA[j] or "")
                if len(parts) > MAX_PARTS:
                    app.logger.warning(f"Chunk starting at row {i+1} too large. Skipping...")
                    break
                j += 1

            if j >= n:
                app.logger.warning(f"Incomplete XML chunk starting at row {i+1}. Skipping...")
                i += 1
                continue

            # Include the closing tag line
            parts.append(colA[j] or "")
            raw_xml = "".join(parts)

            if len(raw_xml.encode("utf-8")) > MAX_XML_BYTES:
                app.logger.warning(f"Chunk starting at row {i+1} exceeds size limit. Skipping...")
                i = j + 1
                continue

            try:
                messages = _extract_messages(raw_xml)
            except Exception as e:
                app.logger.warning(f"Message extraction failed at row {i+1}: {e}")
                i = j + 1
                continue

            for msg in messages:
                current_row = process_xml_and_populate_xl_sheet(msg, ws_out, current_row)

            i = j + 1
        else:
            i += 1

    wb_in.close()
    wb_out.save(xl_output_file)
    wb_out.close()

    return xl_output_file


@app.get("/")
def index():
    return render_template("index.html")


@app.post("/process")
def process_route():
    app.logger.info("Received /process")
    t0 = time.perf_counter()

    if "file" not in request.files:
        return jsonify({"ok": False, "error": "No file part"}), 400

    f = request.files["file"]
    if not f.filename:
        return jsonify({"ok": False, "error": "No selected file"}), 400

    if not _is_allowed(f.filename):
        return jsonify({"ok": False, "error": "Only .xlsx files allowed"}), 400

    orig_name = secure_filename(f.filename)
    tmp_dir = Path(tempfile.mkdtemp(prefix="excelproc_"))
    input_path = tmp_dir / orig_name
    output_path = tmp_dir / f"processed_{orig_name}"

    try:
        f.save(str(input_path))
    except Exception as e:
        return jsonify({"ok": False, "error": f"Save error: {e}"}), 500

    t1 = time.perf_counter()
    app.logger.info(f"Upload saved in {t1 - t0:.2f}s")

    try:
        final_output = process_excel(str(input_path), str(output_path))
    except Exception as e:
        app.logger.exception("Processing error")
        return jsonify({"ok": False, "error": str(e)}), 500

    token = uuid.uuid4().hex
    PROCESS_CACHE[token] = (final_output, tmp_dir)

    resp = {
        "ok": True,
        "download_url": url_for("download", token=token, _external=True),
        "filename": Path(final_output).name
    }
    return jsonify(resp), 200


@app.get("/download/<token>")
def download(token: str):
    if token not in PROCESS_CACHE:
        abort(404, description="Invalid or expired token")
    output_path, tmp_dir = PROCESS_CACHE.pop(token)

    @after_this_request
    def _cleanup(response):
        try:
            for p in tmp_dir.glob("*"):
                p.unlink(missing_ok=True)
            tmp_dir.rmdir()
        except Exception:
            pass
        return response

    return send_file(
        output_path,
        as_attachment=True,
        download_name=Path(output_path).name,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )


if __name__ == "__main__":
    app.run()
#host=APP_HOST, port=APP_PORT, debug=True